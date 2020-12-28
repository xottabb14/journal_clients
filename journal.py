import openpyxl #time, xlwt, xlrd #pip install openpyxl   pip install xlwt pip install xlrd - библиотека для работы с Экселем
from openpyxl.styles import PatternFill#Подключаем стили для ячеек
from openpyxl.styles.borders import Border, Side
import ctypes
from datetime import datetime

filexls = "Журнал.xlsx"

print ("________Журнал учета посещений клиентов by Xottabb14________")

def curdate ():
	now1 = datetime.now()
	if now1.day <=9:
		dday = "0"+str(now1.day)
	else:
		dday = str(now1.day)
	if now1.month <=9:
		mmonth = "0"+str(now1.month)
	else:
		mmonth = str(now1.month)
	if now1.year <=9:
		yyear = "0"+str(now1.year)
	else:
		yyear = str(now1.year)
	dt = dday+"."+mmonth+"."+yyear #текущая дата
	return dt

def curtime ():
	now1 = datetime.now()
	if now1.hour <=9:
		hhour = "0"+str(now1.hour)
	else:
		hhour = str(now1.hour)
	if now1.minute <=9:
		mminute = "0"+str(now1.minute)
	else:
		mminute = str(now1.minute)

	ntime = hhour+":"+mminute #текущее время
	return ntime
curdt = curdate()
firtstime = curtime ()
print ("Текущая дата: "+curdt)
print ("Время прибытия: "+firtstime)
keypause = input("После работы с клиентом нажмите любую клавишу для ввода данных___")
secondtime = curtime ()
print ("Время убытия: "+secondtime)
FIO = input ("Ведите ФИО клиента:_")
ORG = input ("Введите организацию клиента:_")
target = input ("Введите цель визита:_")
primech = input ("Введите примечание, если нужно:_")
if FIO == "":
	FIO = "-"
if ORG == "":
	ORG = "-"
if target == "":
	target = "-"
if primech == "":
	primech = "-"
FIO = FIO+" "+ORG

def readxls (filexls):
	wb = openpyxl.load_workbook(filename = filexls)
	sheet = wb['Коленков'] #выбираем лист
	nums = [v[0].value for v in sheet['A1':'A1000']]

	numlist = []

	for sh in nums:
		if sh != None:
			numlist.append(str(sh))
		else:
			break

	return numlist
numlist = readxls(filexls)

def sheetwrite(numlist,datelist,fiolist,cellist,tinlist,toutlist,primlist):
	wb = openpyxl.load_workbook(filename = filexls)
	sheet = wb['Коленков'] #выбираем лист
	sheet [('A'+str(len(numlist)+1))] = str(len(numlist))
	sheet [('B'+str(len(numlist)+1))] = datelist
	sheet [('C'+str(len(numlist)+1))] = fiolist
	sheet [('D'+str(len(numlist)+1))] = cellist
	sheet [('E'+str(len(numlist)+1))] = tinlist
	sheet [('F'+str(len(numlist)+1))] = toutlist
	sheet [('G'+str(len(numlist)+1))] = primlist

	for s in ["A","C","E","F"]:
		work_sheet = sheet [(s+str(len(numlist)+1))]
		work_sheet.fill = PatternFill(fill_type='solid', start_color='b5e9ff', end_color='b5e9ff')#Данный код позволяет делать оформление цветом ячейки
	for s in ["B","D","G"]:
		work_sheet = sheet [(s+str(len(numlist)+1))]
		work_sheet.fill = PatternFill(fill_type='solid', start_color='ffd88f', end_color='ffd88f')#Данный код позволяет делать оформление цветом ячейки
	thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
	for s in ["A","B","C","D","E","F","G"]:
		work_sheet = sheet [(s+str(len(numlist)+1))]
		work_sheet.border = thin_border
	try:
		wb.save (filexls)
		print ("Журнал посещений сохранен и заполнен.")
	except:
		wb.save (filexls+"_unsaved.xlsx")
		text = "Не удалось сохранить журнал посещений, так как он уже используется. Копия сохранена в файл "+filexls+"_unsaved.xlsx"
		ctypes.windll.user32.MessageBoxW(0, text, "Журнал учета посещений by Xottabb14", 0)


sheetwrite (numlist,curdt,FIO,target,firtstime,secondtime,primech)
keypause = input ("Нажмите любую клавишу для выхода...")